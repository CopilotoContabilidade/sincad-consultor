import io, os, time, base64
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import requests as req
from bs4 import BeautifulSoup
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image, ImageEnhance
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

app = Flask(__name__)

SINCAD_URL  = "https://sucief-sincad-web.fazenda.rj.gov.br/sincad-web/index.jsf"
SINCAD_BASE = "https://sucief-sincad-web.fazenda.rj.gov.br"

# ── Driver local com Chrome headless ──────────────────────────────────────
def criar_driver():
    opts = Options()
    opts.add_argument('--headless=new')
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-dev-shm-usage')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--window-size=1366,768')
    opts.add_argument('--disable-blink-features=AutomationControlled')
    opts.add_experimental_option('excludeSwitches', ['enable-automation'])
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)

# ── CAPTCHA via Claude Vision ──────────────────────────────────────────────
def resolver_captcha(img_bytes):
    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip().strip('"\'')
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY nao configurada no ambiente")
    img = Image.open(io.BytesIO(img_bytes)).convert('L')
    img = ImageEnhance.Contrast(img).enhance(2.5)
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    img = img.convert('RGB')
    buf = io.BytesIO(); img.save(buf, format='PNG')
    img_b64 = base64.standard_b64encode(buf.getvalue()).decode('utf-8')
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=20,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": img_b64}},
            {"type": "text", "text": "CAPTCHA image. Reply ONLY with the characters you see, nothing else."}
        ]}]
    )
    return msg.content[0].text.strip()

# ── Helpers ────────────────────────────────────────────────────────────────
def limpar_cnpj(v): return ''.join(filter(str.isdigit, str(v)))
def fmt_cnpj(v):
    c = limpar_cnpj(v).zfill(14)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"

def preencher(driver, el, valor):
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});"
            "arguments[0].removeAttribute('readonly');"
            "arguments[0].removeAttribute('disabled');", el)
        time.sleep(0.2)
        el.click(); time.sleep(0.1)
        el.clear(); el.send_keys(valor)
        return True
    except: pass
    try:
        driver.execute_script(
            "var e=arguments[0],v=arguments[1];"
            "e.removeAttribute('readonly');e.removeAttribute('disabled');"
            "e.value=v;"
            "['input','change'].forEach(t=>e.dispatchEvent(new Event(t,{bubbles:true})));",
            el, valor)
        return True
    except: return False

# ── Consulta ───────────────────────────────────────────────────────────────
def consultar_cnpj(cnpj_raw, max_tent=4):
    cnpj = limpar_cnpj(cnpj_raw)
    driver = None
    try:
        driver = criar_driver()
        wait = WebDriverWait(driver, 15)

        for t in range(1, max_tent + 1):
            try:
                driver.get(SINCAD_URL)
                time.sleep(3)  # Aguarda JS renderizar o formulário

                # ── Campo CNPJ ──
                campo_cnpj = None
                for xp in [
                    "//input[contains(translate(@id,'CNPJ','cnpj'),'cnpj')]",
                    "//input[contains(translate(@name,'CNPJ','cnpj'),'cnpj')]",
                    "(//input[@type='text'])[1]",
                    "(//input[not(@type='hidden')])[1]",
                ]:
                    try:
                        campo_cnpj = wait.until(EC.presence_of_element_located((By.XPATH, xp)))
                        break
                    except: continue

                if not campo_cnpj:
                    continue

                preencher(driver, campo_cnpj, fmt_cnpj(cnpj))
                time.sleep(0.5)

                # ── CAPTCHA imagem ──
                cap_el = None
                for xp in [
                    "//img[contains(@src,'botdetect')]",
                    "//img[contains(@src,'get=image')]",
                    "//img[contains(@src,'captcha') or contains(@src,'Captcha')]",
                ]:
                    try: cap_el = driver.find_element(By.XPATH, xp); break
                    except: continue

                if not cap_el:
                    continue

                cap_src = cap_el.get_attribute('src')
                if cap_src.startswith('/'): cap_src = SINCAD_BASE + cap_src
                cookies = {c['name']: c['value'] for c in driver.get_cookies()}
                cap_r = req.get(cap_src, cookies=cookies, timeout=10)
                captcha_code = resolver_captcha(cap_r.content)

                # ── Campo CAPTCHA ──
                campo_cap = None
                for xp in [
                    "//input[contains(@class,'BDC_CaptchaInput')]",
                    "//input[contains(@id,'BDC') or contains(@name,'BDC')]",
                    "//input[contains(translate(@id,'CAPTCHA','captcha'),'captcha') and not(contains(translate(@id,'CNPJ','cnpj'),'cnpj'))]",
                    "//input[contains(translate(@name,'CAPTCHA','captcha'),'captcha')]",
                    "(//input[@type='text'])[last()]",
                    "(//input[not(@type='hidden')])[last()]",
                ]:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el != campo_cnpj: campo_cap = el; break
                    except: continue

                if not campo_cap:
                    continue

                preencher(driver, campo_cap, captcha_code)
                time.sleep(0.3)

                # ── Submit ──
                btn = None
                for xp in [
                    "//input[@type='submit']",
                    "//button[@type='submit']",
                    "//button[contains(normalize-space(),'Pesquis')]",
                    "//input[contains(@value,'Pesquis')]",
                ]:
                    try: btn = driver.find_element(By.XPATH, xp); break
                    except: continue

                if not btn: continue
                btn.click()
                time.sleep(4)

                # ── Resultado ──
                body = driver.find_element(By.TAG_NAME, 'body').text.lower()
                erros_cap = ['captcha inválido', 'código inválido', 'invalid captcha',
                             'captcha incorreto', 'tente novamente', 'informe o código']
                if any(x in body for x in erros_cap):
                    continue

                if 'não há registros' in body or 'nenhum registro' in body:
                    return {'condicao': 'Sem registros', 'ie_encontrada': ''}

                soup = BeautifulSoup(driver.page_source, 'html.parser')
                conds, ies = [], []
                for table in soup.find_all('table'):
                    for row in table.find_all('tr'):
                        cells = [td.get_text(strip=True) for td in row.find_all('td')]
                        if not cells: continue
                        if any('condição da inscrição' in c.lower() for c in cells): continue
                        if len(cells) >= 5: ies.append(cells[3]); conds.append(cells[4])
                        elif len(cells) >= 2 and cells[-1]: conds.append(cells[-1])
                if conds:
                    return {'condicao': ' | '.join(filter(None, conds)),
                            'ie_encontrada': ' | '.join(filter(None, ies))}
                if t == max_tent:
                    return {'condicao': 'Resultado não reconhecido', 'ie_encontrada': ''}

            except Exception as e:
                if t == max_tent:
                    return {'condicao': f'Erro: {str(e)[:80]}', 'ie_encontrada': ''}
                time.sleep(1)

    finally:
        if driver:
            try: driver.quit()
            except: pass

    return {'condicao': 'Falha', 'ie_encontrada': ''}

# ── Excel ──────────────────────────────────────────────────────────────────
def gerar_excel(empresas):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SINCAD"
    headers = ['Nome','CNPJ','IE (entrada)','IE (SINCAD)','Situação','Consultado em']
    widths  = [45, 22, 22, 22, 38, 22]
    hf = PatternFill("solid", fgColor="472D54"); ht = Font(bold=True, color="FFFFFF")
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill=hf; c.font=ht; c.alignment=Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    CV=PatternFill("solid",fgColor="C6EFCE"); CR=PatternFill("solid",fgColor="FFC7CE")
    CA=PatternFill("solid",fgColor="FFEB9C"); CC=PatternFill("solid",fgColor="EEEEEE")
    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    for i, e in enumerate(empresas, 2):
        cl = e.get('condicao','').lower()
        if any(p in cl for p in ['habilitad','ativa','regular']): fill=CV
        elif any(p in cl for p in ['baix','cancel','inapt','irregular']): fill=CR
        elif any(p in cl for p in ['suspens','sem registro']): fill=CA
        else: fill=CC
        vals = [e.get('nome',''), fmt_cnpj(e.get('cnpj','')),
                e.get('ie',''), e.get('ie_encontrada',''), e.get('condicao',''), now]
        for col, val in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=val).fill = fill
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── Rotas ──────────────────────────────────────────────────────────────────
@app.route('/')
def index(): return HTML

@app.route('/api/consultar', methods=['POST'])
def api_consultar():
    data = request.get_json() or {}
    cnpj = data.get('cnpj','').strip()
    if not cnpj: return jsonify({'erro':'CNPJ não informado'}), 400
    res = consultar_cnpj(cnpj)
    res['cnpj_fmt'] = fmt_cnpj(cnpj)
    return jsonify(res)

@app.route('/api/download', methods=['POST'])
def api_download():
    empresas = request.get_json() or []
    buf = gerar_excel(empresas)
    fname = f"resultado_sincad_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Consultor SINCAD — Copiloto</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,-apple-system,sans-serif;background:#f5f0eb;color:#303030;min-height:100vh}
.hdr{background:#472d54;padding:20px 28px;display:flex;align-items:center;justify-content:space-between}
.hdr h1{color:#fff;font-size:18px;font-weight:600}.hdr p{color:#c4afd4;font-size:12px;margin-top:2px}
.badge{background:#c17f3e;color:#fff;font-size:12px;font-weight:600;padding:4px 14px;border-radius:20px}
.main{max-width:960px;margin:28px auto;padding:0 20px}
.card{background:#fff;border-radius:12px;padding:22px;margin-bottom:18px;border:1px solid #e8e0d8}
.card h2{font-size:13px;font-weight:700;color:#472d54;text-transform:uppercase;letter-spacing:.06em;margin-bottom:16px}
.fg{display:grid;grid-template-columns:2fr 1.4fr 1.4fr auto;gap:10px;align-items:end;margin-bottom:14px}
label.lbl{font-size:11px;color:#888;display:block;margin-bottom:4px;text-transform:uppercase;letter-spacing:.04em}
input[type=text]{width:100%;padding:9px 11px;border:1px solid #ddd;border-radius:8px;font-size:13px;color:#303030;font-family:inherit}
input[type=text]:focus{outline:none;border-color:#472d54}
.btn{padding:9px 18px;border-radius:8px;font-size:13px;font-weight:600;cursor:pointer;border:none;font-family:inherit;white-space:nowrap}
.bp{background:#472d54;color:#fff}.bp:hover{background:#5d3d6e}.bp:disabled{opacity:.4;cursor:not-allowed}
.bs{background:#fff;color:#472d54;border:1px solid #472d54}.bs:hover{background:#f8f4fb}
.bg{background:#2e7d32;color:#fff}.bg:hover{background:#1b5e20}
.bd{background:none;border:none;color:#ccc;font-size:18px;cursor:pointer;padding:2px 8px;border-radius:4px}
.bd:hover{color:#b24a4a;background:#fff0f0}
.tbar{display:flex;gap:8px;margin-bottom:14px;align-items:center;flex-wrap:wrap}
.src{flex:1;min-width:120px;padding:8px 12px;border:1px solid #ddd;border-radius:8px;font-size:13px;font-family:inherit}
.src:focus{outline:none;border-color:#472d54}
.tw{border:1px solid #ede5da;border-radius:10px;overflow:hidden}
.thr{display:grid;grid-template-columns:2fr 1.4fr 1.4fr 36px;padding:9px 14px;background:#472d54;gap:10px}
.th{color:#c4afd4;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.06em}
.tr{display:grid;grid-template-columns:2fr 1.4fr 1.4fr 36px;padding:11px 14px;gap:10px;align-items:center;border-bottom:1px solid #f0e8e0}
.tr:last-child{border-bottom:none}.tr:nth-child(even){background:#faf8f5}
.tn{font-size:13px;font-weight:500;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.tm{font-size:12px;color:#555;font-family:monospace}.td{font-size:12px;color:#888}
.empty{padding:36px;text-align:center;color:#aaa;font-size:13px}
.pb{background:#f0e8e0;border-radius:8px;height:10px;overflow:hidden;margin-bottom:6px}
.pf{background:#472d54;height:10px;border-radius:8px;transition:width .3s}.pt{font-size:12px;color:#888}
.rt{width:100%;border-collapse:collapse;font-size:12px}
.rt th{background:#472d54;color:#fff;padding:8px 12px;text-align:left;font-size:11px;letter-spacing:.04em}
.rt td{padding:8px 12px;border-bottom:1px solid #f0e8e0}
.rt tr:last-child td{border-bottom:none}
.tag{display:inline-block;padding:2px 9px;border-radius:12px;font-size:11px;font-weight:600}
.tok{background:#e8f5e9;color:#2e7d32}.tbd{background:#fce8e8;color:#b24a4a}
.twr{background:#fff8e1;color:#f57f17}.tgy{background:#f5f5f5;color:#757575}.tpd{background:#e3f2fd;color:#1565c0}
.toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%);background:#303030;color:#fff;padding:10px 20px;border-radius:20px;font-size:13px;z-index:999;opacity:0;transition:opacity .3s;pointer-events:none}
.info{background:#fef9f0;border-left:3px solid #c17f3e;border-radius:0 8px 8px 0;padding:10px 14px;font-size:12px;color:#7a5c1e;margin-top:14px}
</style>
</head>
<body>
<div class="hdr">
  <div><h1>Consultor SINCAD — SEFAZ-RJ</h1><p>Copiloto Contabilidade</p></div>
  <span class="badge" id="badge">0 empresas</span>
</div>
<div class="main">
  <div class="card">
    <h2>Empresas para consulta</h2>
    <div class="fg">
      <div><label class="lbl">Nome *</label><input type="text" id="fn" placeholder="Ex: Tech Ltda"></div>
      <div><label class="lbl">CNPJ *</label><input type="text" id="fc" placeholder="00.000.000/0000-00" maxlength="18"></div>
      <div><label class="lbl">IE</label><input type="text" id="fi" placeholder="Opcional"></div>
      <div><label class="lbl">&nbsp;</label><button class="btn bp" onclick="add()">+ Adicionar</button></div>
    </div>
    <div class="tbar">
      <label class="btn bs" style="cursor:pointer">📥 Importar planilha<input type="file" accept=".xlsx,.xls" onchange="imp(event)" style="display:none"></label>
      <button class="btn bp" id="btnc" onclick="start()" disabled>▶ Consultar tudo</button>
      <input class="src" placeholder="🔍 Buscar..." oninput="flt(this.value)">
    </div>
    <div class="tw">
      <div class="thr"><span class="th">Nome</span><span class="th">CNPJ</span><span class="th">IE</span><span class="th"></span></div>
      <div id="lst"><div class="empty">Adicione empresas ou importe um .xlsx</div></div>
    </div>
    <div class="info">📋 Coluna A = Nome &nbsp;|&nbsp; B = CNPJ &nbsp;|&nbsp; C = IE (opcional)</div>
  </div>
  <div class="card" id="cres" style="display:none">
    <h2>Resultados</h2>
    <div id="prog"></div>
    <div style="display:flex;gap:8px;margin-bottom:14px">
      <button class="btn bg" id="btndl" onclick="dl()" style="display:none">📥 Baixar Excel</button>
      <button class="btn bs" id="btnst" onclick="stop()">⏹ Parar</button>
    </div>
    <div style="overflow-x:auto">
      <table class="rt"><thead><tr><th>Nome</th><th>CNPJ</th><th>Situação</th><th>IE encontrada</th></tr></thead>
      <tbody id="rbody"></tbody></table>
    </div>
  </div>
</div>
<div class="toast" id="toast"></div>
<script>
const K='copiloto-sincad';
let emps=[],flt_='',run=false,stp=false,res=[];
try{emps=JSON.parse(localStorage.getItem(K)||'[]')}catch(e){}
function sv(){try{localStorage.setItem(K,JSON.stringify(emps))}catch(e){}}
function toast(m,d=2500){const t=document.getElementById('toast');t.textContent=m;t.style.opacity='1';setTimeout(()=>t.style.opacity='0',d)}
function fc(s){const d=(s||'').replace(/\D/g,'').padStart(14,'0').slice(0,14);return d.slice(0,2)+'.'+d.slice(2,5)+'.'+d.slice(5,8)+'/'+d.slice(8,12)+'-'+d.slice(12)}
function icnpj(e){const d=e.target.value.replace(/\D/g,'').slice(0,14);let f=d;if(d.length>12)f=d.slice(0,2)+'.'+d.slice(2,5)+'.'+d.slice(5,8)+'/'+d.slice(8,12)+'-'+d.slice(12);else if(d.length>8)f=d.slice(0,2)+'.'+d.slice(2,5)+'.'+d.slice(5,8)+'/'+d.slice(8);else if(d.length>5)f=d.slice(0,2)+'.'+d.slice(2,5)+'.'+d.slice(5);else if(d.length>2)f=d.slice(0,2)+'.'+d.slice(2);e.target.value=f}
document.getElementById('fc').addEventListener('input',icnpj);
['fn','fc','fi'].forEach(id=>document.getElementById(id).addEventListener('keydown',e=>{if(e.key==='Enter')add()}));
function add(){
  const n=document.getElementById('fn').value.trim(),c=document.getElementById('fc').value.replace(/\D/g,''),ie=document.getElementById('fi').value.trim();
  if(!n){toast('Informe o nome');return}if(c.length!==14){toast('CNPJ: 14 dígitos');return}
  if(emps.some(e=>e.cnpj===c)){toast('CNPJ já na lista');return}
  emps.push({id:Date.now(),nome:n,cnpj:c,ie});sv();render();
  ['fn','fc','fi'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('fn').focus();toast(n+' adicionada')}
function rem(id){emps=emps.filter(e=>e.id!==id);sv();render()}
function flt(q){flt_=q.toLowerCase();render()}
function render(){
  const lst=document.getElementById('lst'),bdg=document.getElementById('badge'),btn=document.getElementById('btnc');
  bdg.textContent=emps.length+' empresa'+(emps.length!==1?'s':'');btn.disabled=emps.length===0||run;
  const v=flt_?emps.filter(e=>e.nome.toLowerCase().includes(flt_)||e.cnpj.includes(flt_.replace(/\D/g,''))):emps;
  if(!v.length){lst.innerHTML='<div class="empty">'+(emps.length===0?'Nenhuma empresa ainda':'Nenhuma encontrada')+'</div>';return}
  lst.innerHTML=v.map(e=>`<div class="tr"><span class="tn" title="${e.nome}">${e.nome}</span><span class="tm">${fc(e.cnpj)}</span><span class="td">${e.ie||'—'}</span><button class="bd" onclick="rem(${e.id})">×</button></div>`).join('')}
function imp(e){
  const f=e.target.files[0];if(!f)return;
  const r=new FileReader();
  r.onload=ev=>{try{
    const wb=XLSX.read(ev.target.result,{type:'binary'}),ws=wb.Sheets[wb.SheetNames[0]];
    const rows=XLSX.utils.sheet_to_json(ws,{header:1,defval:''});
    const ns=rows.slice(1).map(r=>({nome:String(r[0]||'').trim(),cnpj:String(r[1]||'').replace(/\D/g,''),ie:String(r[2]||'').trim()})).filter(r=>r.cnpj.length===14&&r.nome);
    if(!ns.length){toast('Nenhuma válida');return}
    const ex=new Set(emps.map(e=>e.cnpj));
    const add=ns.filter(n=>!ex.has(n.cnpj)).map(n=>({...n,id:Date.now()+Math.random()}));
    emps=[...emps,...add];sv();render();toast(add.length+' importada(s)')
  }catch(err){toast('Erro ao ler arquivo')}};
  r.readAsBinaryString(f);e.target.value=''}
function stag(c){const cl=(c||'').toLowerCase();
  if(cl.includes('habilitad')||cl.includes('ativa')||cl.includes('regular'))return`<span class="tag tok">${c}</span>`;
  if(cl.includes('baix')||cl.includes('cancel')||cl.includes('inapt'))return`<span class="tag tbd">${c}</span>`;
  if(cl.includes('suspens')||cl.includes('sem registro'))return`<span class="tag twr">${c}</span>`;
  if(cl.includes('erro')||cl.includes('falha'))return`<span class="tag tgy">${c}</span>`;
  return`<span class="tag tpd">${c||'Aguardando'}</span>`}
function prog(f,t){document.getElementById('prog').innerHTML=`<div class="pb"><div class="pf" style="width:${Math.round(f/t*100)}%"></div></div><div class="pt">${f} de ${t} — cada empresa leva ~30 segundos</div>`}
async function start(){
  if(!emps.length)return;run=true;stp=false;res=[];
  document.getElementById('cres').style.display='block';
  document.getElementById('rbody').innerHTML='';
  document.getElementById('btndl').style.display='none';
  document.getElementById('btnc').disabled=true;
  document.getElementById('btnst').style.display='inline-block';
  prog(0,emps.length);
  window.scrollTo({top:document.getElementById('cres').offsetTop-20,behavior:'smooth'});
  for(let i=0;i<emps.length;i++){
    if(stp)break;
    const emp=emps[i];let r;
    try{
      const resp=await fetch('/api/consultar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cnpj:emp.cnpj})});
      r=await resp.json();
    }catch(err){r={condicao:'Erro de conexão',ie_encontrada:''}}
    res.push({...emp,...r,cnpj_fmt:fc(emp.cnpj)});
    const tr=document.createElement('tr');
    tr.innerHTML=`<td>${emp.nome}</td><td style="font-family:monospace;font-size:11px">${fc(emp.cnpj)}</td><td>${stag(r.condicao)}</td><td style="font-size:11px;color:#888">${r.ie_encontrada||'—'}</td>`;
    document.getElementById('rbody').prepend(tr);prog(i+1,emps.length)}
  run=false;document.getElementById('btnc').disabled=false;
  document.getElementById('btnst').style.display='none';
  if(res.length)document.getElementById('btndl').style.display='inline-block';
  toast('Consulta finalizada!'+(stp?' (interrompida)':''))}
function stop(){stp=true;document.getElementById('btnst').style.display='none';toast('Parando...')}
async function dl(){
  const r=await fetch('/api/download',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(res)});
  const b=await r.blob(),u=URL.createObjectURL(b),a=document.createElement('a');
  a.href=u;a.download=`resultado_sincad_${new Date().toISOString().slice(0,10)}.xlsx`;a.click();URL.revokeObjectURL(u)}
render();
</script></body></html>"""

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
